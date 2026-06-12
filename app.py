import os
import re
import base64
import pdfplumber
import openpyxl
import msoffcrypto
import requests
from flask import Flask, request, jsonify
from io import BytesIO

app = Flask(__name__)

SUPABASE_URL = "https://kjtohbkajcscayckyypi.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtqdG9oYmthamNzY2F5Y2t5eXBpIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODA5OTkzMTgsImV4cCI6MjA5NjU3NTMxOH0.VzKnzpf4VFi06kn28Wz_b8kcWwujF7d5lQG0Xwa6aiw"

PDF_PASSWORDS = ["0812", "9891"]

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates"
}

def decode_pdf(raw_body):
    """Handle both raw binary and base64-encoded PDF from Power Automate."""
    if isinstance(raw_body, bytes):
        try:
            decoded = base64.b64decode(raw_body)
            if decoded[:4] == b'%PDF':
                return decoded
        except Exception:
            pass
        if raw_body[:4] == b'%PDF':
            return raw_body
        try:
            text = raw_body.decode('utf-8').strip().strip('"')
            decoded = base64.b64decode(text)
            if decoded[:4] == b'%PDF':
                return decoded
        except Exception:
            pass
    return raw_body

def open_pdf(pdf_bytes):
    """Try opening PDF with known passwords, fall back to no password."""
    for pwd in PDF_PASSWORDS:
        try:
            return pdfplumber.open(BytesIO(pdf_bytes), password=pwd)
        except Exception:
            continue
    try:
        return pdfplumber.open(BytesIO(pdf_bytes))
    except Exception as e:
        raise Exception(f"Could not open PDF with any known password: {str(e)}")

def detect_company(source_file, full_text=""):
    """Detect company from source filename/subject or PDF text."""
    combined = (source_file + " " + full_text).upper()
    if "INFOCOM" in combined:
        return "Infocom"
    if "INSPIRO" in combined:
        return "Inspiro"
    return "Inspiro"  # default

def extract_approved_date(full_text):
    """
    Extract the approval date from text like:
    'Successful Payroll Accounts Opened on 06/10/2026 - 06/11/2026 via Digital Payroll Portal'
    Returns the LAST (end) date in MM/DD/YYYY -> YYYY-MM-DD format.
    """
    pattern = r"Opened on\s+(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})"
    match = re.search(pattern, full_text, re.IGNORECASE)
    if match:
        end_date = match.group(2)
        try:
            mm, dd, yyyy = end_date.split("/")
            return f"{yyyy}-{int(mm):02d}-{int(dd):02d}"
        except Exception:
            return end_date
    # Fallback: any single date pattern "Opened on MM/DD/YYYY"
    pattern2 = r"Opened on\s+(\d{1,2}/\d{1,2}/\d{4})"
    match2 = re.search(pattern2, full_text, re.IGNORECASE)
    if match2:
        d = match2.group(1)
        try:
            mm, dd, yyyy = d.split("/")
            return f"{yyyy}-{int(mm):02d}-{int(dd):02d}"
        except Exception:
            return d
    return ""

def normalize_reason(r):
    r = r.strip().upper()
    if "NO HR" in r or "KYC" in r:
        return "No HR/KYC Certification"
    if "MOBILE" in r:
        return "Mismatch: Mobile Number"
    if "BIRTH" in r or "DOB" in r:
        return "Mismatch: Date of Birth"
    if "MATCHED" in r:
        return "HR Cert. Already Matched"
    return r.title()

def parse_pending(pdf_bytes, source_file):
    records = []
    with open_pdf(pdf_bytes) as pdf:
        full_text = "\n".join((page.extract_text() or "") for page in pdf.pages)
        company = detect_company(source_file, full_text)

        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or not row[0]:
                    continue
                try:
                    no = int(str(row[0]).strip())
                except Exception:
                    continue

                first    = str(row[1] or "").strip()
                middle   = str(row[2] or "").strip()
                last     = str(row[3] or "").strip()
                suffix   = str(row[4] or "").strip()
                app_date = str(row[5] or "").strip()
                dob      = str(row[6] or "").strip()
                mobile   = str(row[7] or "").strip()
                reason   = str(row[8] or "").strip()

                parts = [p for p in [first, middle, last, suffix] if p]
                full_name = " ".join(parts)

                if app_date and " " in app_date:
                    app_date = app_date.split(" ")[0]

                records.append({
                    "row_no": no,
                    "full_name": full_name,
                    "date_of_birth": dob,
                    "application_date": app_date,
                    "mobile_number": mobile,
                    "reason": normalize_reason(reason) if reason else "",
                    "source_file": source_file,
                    "company": company
                })
    return records

def parse_approved(pdf_bytes, source_file, fallback_date=""):
    records = []
    with open_pdf(pdf_bytes) as pdf:
        full_text = "\n".join((page.extract_text() or "") for page in pdf.pages)
        company = detect_company(source_file, full_text)
        date_approved = extract_approved_date(full_text) or fallback_date

        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or not row[0]:
                    continue
                try:
                    no = int(str(row[0]).strip())
                except Exception:
                    continue

                first   = str(row[1] or "").strip()
                middle  = str(row[2] or "").strip()
                last    = str(row[3] or "").strip()
                suffix  = str(row[4] or "").strip()
                account = str(row[5] or "").strip()
                branch  = str(row[6] or "").strip()

                parts = [p for p in [first, middle, last, suffix] if p]
                full_name = " ".join(parts)

                records.append({
                    "row_no": no,
                    "full_name": full_name,
                    "account_number": account,
                    "rcbc_branch": branch,
                    "date_approved": date_approved,
                    "source_file": source_file,
                    "company": company
                })
    return records

XLSX_PASSWORDS = ["0812", "9891"]

def decrypt_xlsx(file_bytes):
    """Try to decrypt password-protected xlsx; return plain bytes."""
    for pwd in XLSX_PASSWORDS:
        try:
            office_file = msoffcrypto.OfficeFile(BytesIO(file_bytes))
            office_file.load_key(password=pwd)
            decrypted = BytesIO()
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            data = decrypted.read()
            if data[:2] == b'PK':
                return data
        except Exception:
            continue
    # Not encrypted or no password worked - return as-is
    return file_bytes

def parse_approved_xlsx(file_bytes, source_file):
    """Parse the cumulative Excel master list of approved accounts."""
    records = []
    file_bytes = decrypt_xlsx(file_bytes)
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active

    company = detect_company(source_file)

    # Find header row (assume row 1) and map columns
    headers = [str(c.value or "").strip().lower() for c in ws[1]]

    def col_idx(*names):
        for name in names:
            for i, h in enumerate(headers):
                if name in h:
                    return i
        return None

    idx_first  = col_idx("first name")
    idx_middle = col_idx("middle name")
    idx_last   = col_idx("last name")
    idx_suffix = col_idx("suffix")
    idx_acct   = col_idx("account number", "account no")
    idx_branch = col_idx("branch")
    idx_date   = col_idx("disposition date", "date approved", "date")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if idx_acct is None or idx_acct >= len(row):
            continue
        account = str(row[idx_acct] or "").strip()
        if not account:
            continue

        first  = str(row[idx_first] or "").strip()  if idx_first  is not None else ""
        middle = str(row[idx_middle] or "").strip() if idx_middle is not None else ""
        last   = str(row[idx_last] or "").strip()   if idx_last   is not None else ""
        suffix = str(row[idx_suffix] or "").strip() if idx_suffix is not None else ""
        branch = str(row[idx_branch] or "").strip() if idx_branch is not None else ""

        date_val = row[idx_date] if idx_date is not None and idx_date < len(row) else ""
        if hasattr(date_val, "strftime"):
            date_approved = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val or "").strip()
            # Try MM-DD-YYYY or MM/DD/YYYY -> YYYY-MM-DD
            m = re.match(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", date_str)
            if m:
                mm, dd, yyyy = m.groups()
                date_approved = f"{yyyy}-{int(mm):02d}-{int(dd):02d}"
            else:
                date_approved = date_str

        parts = [p for p in [first, middle, last, suffix] if p]
        full_name = " ".join(parts)

        records.append({
            "full_name": full_name,
            "account_number": account,
            "rcbc_branch": branch,
            "date_approved": date_approved,
            "source_file": source_file,
            "company": company
        })
    return records

def save_to_supabase(table, records, on_conflict=None, ignore_duplicates=False):
    if not records:
        return {"saved": 0, "message": "No records parsed"}
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if on_conflict:
        url += f"?on_conflict={on_conflict}"
    headers = dict(HEADERS)
    if ignore_duplicates:
        headers["Prefer"] = "resolution=ignore-duplicates"
    res = requests.post(
        url,
        headers=headers,
        json=records
    )
    return {"saved": len(records), "status": res.status_code, "response": res.text}

@app.route("/parse-pending", methods=["POST"])
def handle_pending():
    try:
        raw = request.data
        pdf_bytes = decode_pdf(raw)
        source_file = request.headers.get("X-Source-File", "unknown")
        records = parse_pending(pdf_bytes, source_file)
        result = save_to_supabase("pending", records)
        return jsonify(result), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/parse-approved", methods=["POST"])
def handle_approved():
    try:
        raw = request.data
        source_file = request.headers.get("X-Source-File", "unknown")

        # Detect xlsx by magic bytes (PK = zip-based office format) or raw not matching %PDF
        file_bytes = raw
        try:
            decoded = base64.b64decode(raw)
            file_bytes = decoded
        except Exception:
            file_bytes = raw

        if file_bytes[:2] == b'PK':
            # Excel file (.xlsx is a zip archive)
            records = parse_approved_xlsx(file_bytes, source_file)
            result = save_to_supabase("approved", records, on_conflict="account_number", ignore_duplicates=True)
        else:
            # Fallback: legacy PDF support
            pdf_bytes = decode_pdf(raw)
            fallback_date = request.headers.get("X-Date-Approved", "")
            records = parse_approved(pdf_bytes, source_file, fallback_date)
            result = save_to_supabase("approved", records, on_conflict="account_number", ignore_duplicates=True)

        return jsonify(result), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/debug-pdf", methods=["POST"])
def debug_pdf():
    try:
        raw = request.data
        pdf_bytes = decode_pdf(raw)
        with open_pdf(pdf_bytes) as pdf:
            pages_text = []
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or "(no text extracted)"
                pages_text.append({"page": i+1, "text": text[:1000]})
        return jsonify({"pages": pages_text}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
